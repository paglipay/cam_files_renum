"""
cam_mapper.py  —  local Streamlit app for camera GPS mapping and renaming.

Workflow:
  1. Scan folder for *_INSTALL* images, extract EXIF GPS.
  2. View cameras on the satellite map.  Hover a pin for the photo preview.
     Click "Get List" (bottom-right of map) to copy all camera numbers,
     paste into the Sequence box, reorder them, then click Apply.
  3. Save cam_mapping.json (same folder as rename_files.xlsx).
  4. Apply JSON → update 'replace' column in rename_files.xlsx input sheet.
  5. Run copy / move from the app.
"""
from __future__ import annotations

import base64, json, os, re, shutil
from datetime import datetime
import streamlit.components.v1 as components
from io import BytesIO
from pathlib import Path

import folium
import openpyxl
import pandas as pd
import piexif
import streamlit as st
from PIL import Image

# ── constants ─────────────────────────────────────────────────────────────────
SATELLITE_TILES = (
    "https://server.arcgisonline.com/ArcGIS/rest/services"
    "/World_Imagery/MapServer/tile/{z}/{y}/{x}"
)
SATELLITE_ATTR = "Esri"
INSTALL_RE     = re.compile(r"_INSTALL", re.IGNORECASE)
ENUMBER_RE     = re.compile(r"(\d+)((?:_[A-Za-z]+|[A-Za-z]+)*)$")
IMAGE_EXTS     = {".jpg", ".jpeg", ".png", ".tif", ".tiff"}
DEFAULT_FOLDER = (
    r"C:\Users\Paul\OneDrive - Los Angeles Unified School District"
    r"\Paul's Share Folder\Project Sites\WILMINGTON PARK EL (7781)"
    r"\Camera\Design\Pictures"
)

st.set_page_config(layout="wide", page_title="Camera Mapper")


# ── Get List button injected into the rendered HTML ───────────────────────────

def _inject_get_list_button(html: str, map_var: str) -> str:
    """
    Append a 'Get List' Leaflet control to the fully-rendered folium HTML.
    Injecting into the HTML string (rather than into folium's render pipeline)
    guarantees the Leaflet map variable is already defined when the script runs.
    The button appears bottom-right on the map.  Clicking it shows an overlay
    with all pin enumbers as a comma-separated list the user can copy and paste
    into the Sequence box below the map.
    """
    js = f"""
<script>
(function() {{
  /* overlay popup */
  var overlay = document.createElement('div');
  overlay.style.cssText = [
    'display:none', 'position:absolute', 'bottom:55px', 'right:10px',
    'z-index:9999', 'background:#fff', 'border:2px solid #555',
    'border-radius:6px', 'padding:12px 14px',
    'box-shadow:0 4px 14px rgba(0,0,0,.45)', 'min-width:280px',
    'font-family:sans-serif'
  ].join(';') + ';';
  overlay.innerHTML = (
    '<div style="font-size:11px;color:#555;margin-bottom:6px;">'
    + 'Copy, paste into the <b>Sequence</b> box below the map, '
    + 'reorder (first = new Cam #1), then click <b>Apply sequence</b>.</div>'
    + '<input id="_camSeqInput" readonly '
    + 'style="width:100%;font-size:14px;font-weight:bold;padding:4px;'
    + 'border:1px solid #aaa;border-radius:3px;box-sizing:border-box;">'
    + '<div style="margin-top:8px;display:flex;justify-content:flex-end;gap:6px;">'
    + '<button id="_camSeqCopy" style="padding:3px 10px;cursor:pointer;'
    + 'background:#1a73e8;color:#fff;border:none;border-radius:3px;font-size:12px;">'
    + '&#128203; Copy</button>'
    + '<button id="_camSeqClose" style="padding:3px 10px;cursor:pointer;'
    + 'border:1px solid #ccc;border-radius:3px;background:#fff;font-size:12px;">'
    + 'Close</button></div>'
  );
  var mc = {map_var}.getContainer();
  mc.style.position = 'relative';
  mc.appendChild(overlay);

  document.getElementById('_camSeqClose').addEventListener('click', function() {{
    overlay.style.display = 'none';
  }});
  document.getElementById('_camSeqCopy').addEventListener('click', function() {{
    var inp = document.getElementById('_camSeqInput');
    inp.select();
    var cb = document.getElementById('_camSeqCopy');
    function done() {{ cb.textContent='Copied!'; setTimeout(function(){{cb.innerHTML='&#128203; Copy';}},1500); }}
    if (navigator.clipboard) {{ navigator.clipboard.writeText(inp.value).then(done); }}
    else {{ document.execCommand('copy'); done(); }}
  }});

  /* Leaflet control */
  var Ctrl = L.Control.extend({{
    options: {{position:'bottomright'}},
    onAdd: function() {{
      var div = L.DomUtil.create('div','leaflet-bar');
      div.style.cssText = 'box-shadow:0 1px 5px rgba(0,0,0,.4);border-radius:4px;';
      var btn = L.DomUtil.create('button','',div);
      btn.innerHTML = '&#128203;&nbsp;Get&nbsp;List';
      btn.title = 'Copy camera numbers to assign the sequence';
      btn.style.cssText = 'padding:6px 12px;cursor:pointer;background:#fff;'
        + 'font-size:13px;font-weight:bold;border:none;border-radius:4px;'
        + 'white-space:nowrap;line-height:1.4;';
      L.DomEvent.on(btn,'click',function(e) {{
        L.DomEvent.stopPropagation(e);
        var nums=[];
        document.querySelectorAll('[data-pin-idx]').forEach(function(el) {{
          var n=parseInt(el.getAttribute('data-pin-idx'));
          if(!isNaN(n)) nums.push(n);
        }});
        nums.sort(function(a,b){{return a-b;}});
        document.getElementById('_camSeqInput').value=nums.join(',');
        overlay.style.display='block';
        setTimeout(function(){{document.getElementById('_camSeqInput').select();}},60);
      }});
      return div;
    }}
  }});
  new Ctrl().addTo({map_var});
}})();
</script>
"""
    # Insert right before </body> — map variable is guaranteed to exist by then
    return html.replace("</body>", js + "\n</body>", 1)


# ── EXIF / scan helpers ───────────────────────────────────────────────────────

def _get_lat_lon(path: str):
    try:
        img        = Image.open(path)
        exif_bytes = img.info.get("exif", b"")
        if not exif_bytes:
            return None, None
        gps = piexif.load(exif_bytes).get("GPS", {})
        if not gps:
            return None, None

        def _deg(v):
            a, b, c = v
            return a[0] / a[1] + b[0] / b[1] / 60 + c[0] / c[1] / 3600

        lat  = gps.get(piexif.GPSIFD.GPSLatitude)
        lref = gps.get(piexif.GPSIFD.GPSLatitudeRef)
        lon  = gps.get(piexif.GPSIFD.GPSLongitude)
        lnrf = gps.get(piexif.GPSIFD.GPSLongitudeRef)
        if lat and lref and lon and lnrf:
            return (
                _deg(lat) * (-1 if lref == b"S" else 1),
                _deg(lon) * (-1 if lnrf == b"W" else 1),
            )
    except Exception:
        pass
    return None, None


def _enumber(stem: str):
    m = ENUMBER_RE.search(stem)
    return (int(m.group(1)), m.group(1), m.group(2)) if m else (None, None, None)


def _thumb(path: str, w: int = 280) -> str:
    try:
        img = Image.open(path)
        img.thumbnail((350, 350))
        buf = BytesIO()
        img.save(buf, "JPEG", quality=80)
        img.close()
        b64 = base64.b64encode(buf.getvalue()).decode()
        return f'<img src="data:image/jpeg;base64,{b64}" width="{w}">'
    except Exception:
        return "<i>preview unavailable</i>"


def scan_folder(folder: str) -> list[dict]:
    results = []
    try:
        entries = sorted(Path(folder).iterdir(), key=lambda e: e.name.lower())
    except OSError as exc:
        st.error(f"Cannot open folder: {exc}")
        return results
    for e in entries:
        if not e.is_file() or e.suffix.lower() not in IMAGE_EXTS:
            continue
        if not INSTALL_RE.search(e.stem):
            continue
        enum, estr, esuf = _enumber(e.stem)
        lat, lon = _get_lat_lon(str(e))
        results.append({
            "path": str(e), "file": e.name,
            "enumber": enum, "enumber_str": estr, "esuffix": esuf,
            "lat": lat, "lon": lon,
        })
    return results


# ── map builder ───────────────────────────────────────────────────────────────

def build_map(images: list[dict], sequence: list[str]) -> folium.Map | None:
    """
    Satellite folium map with numbered DivIcon pins.
    Red = unassigned (shows current enumber).
    Green = assigned (shows proposed sequence position).
    Hover shows a large photo thumbnail.
    'Get List' button (bottom-right) outputs all camera numbers for copy/paste.
    Markers added directly to map (not via FeatureGroup) for reliable DivIcon rendering.
    """
    gps = [r for r in images if r["lat"] is not None]
    if not gps:
        return None

    seq_pos = {fname: i + 1 for i, fname in enumerate(sequence)}
    avg_lat = sum(r["lat"] for r in gps) / len(gps)
    avg_lon = sum(r["lon"] for r in gps) / len(gps)

    m = folium.Map(location=[avg_lat, avg_lon], zoom_start=18, tiles=None)
    folium.TileLayer(
        tiles=SATELLITE_TILES, attr=SATELLITE_ATTR,
        name="Satellite", overlay=False, control=True,
    ).add_to(m)
    folium.TileLayer(
        "OpenStreetMap", name="Street Map", overlay=False, control=True,
    ).add_to(m)
    folium.LayerControl(position="topright", collapsed=False).add_to(m)

    for row in gps:
        pos     = seq_pos.get(row["file"])
        enumber = row["enumber"] if row["enumber"] is not None else "?"
        label   = str(pos) if pos else str(enumber)
        bg      = "#27ae60" if pos else "crimson"
        border  = "3px solid #f1c40f" if pos else "2px solid white"

        thumb = _thumb(row["path"])
        tooltip_html = (
            f'<div style="font-family:sans-serif;max-width:300px">'
            f'<b style="font-size:13px">Cam {enumber}'
            + (f' &rarr; <span style="color:#27ae60">#{pos}</span>' if pos else "")
            + f"</b><br><small>{row['file']}</small>"
            + f'<div style="margin-top:5px">{thumb}</div>'
            + "</div>"
        )

        folium.Marker(
            location=[row["lat"], row["lon"]],
            tooltip=folium.Tooltip(tooltip_html, sticky=False),
            icon=folium.DivIcon(
                html=(
                    f'<div data-pin-idx="{enumber}" style="background:{bg};color:white;'
                    f'font-weight:bold;font-size:12px;width:28px;height:28px;line-height:28px;'
                    f'text-align:center;border-radius:50%;border:{border};'
                    f'box-shadow:0 2px 5px rgba(0,0,0,.6)">{label}</div>'
                ),
                icon_size=(28, 28),
                icon_anchor=(14, 14),
            ),
        ).add_to(m)

    return m


# ── JSON helpers ──────────────────────────────────────────────────────────────

def _json_path(xlsx: str) -> str:
    return str(Path(xlsx).parent / "cam_mapping.json") if xlsx else ""


def save_json(xlsx: str, folder: str, images: list[dict],
              sequence: list[str], non_gps_proposed: dict) -> str:
    seq_pos = {f: i + 1 for i, f in enumerate(sequence)}
    mapping = []
    for r in images:
        if r["enumber"] is None:
            continue
        if r["lat"] is not None:
            proposed = seq_pos.get(r["file"], r["enumber"])
        else:
            proposed = non_gps_proposed.get(r["file"], r["enumber"])
        mapping.append({
            "file":        r["file"],
            "enumber":     r["enumber"],
            "enumber_str": r["enumber_str"],
            "proposed":    proposed,
        })
    data = {
        "folder":    folder,
        "generated": datetime.now().isoformat(timespec="seconds"),
        "mapping":   mapping,
    }
    out = _json_path(xlsx)
    Path(out).write_text(json.dumps(data, indent=2))
    return out


# ── xlsx helpers ──────────────────────────────────────────────────────────────

def _col(ws, name: str) -> int:
    for c in ws[1]:
        if c.value and str(c.value).strip().lower() == name.lower():
            return c.column
    raise ValueError(f"Column '{name}' not found in sheet '{ws.title}'")


def _eff(cell, vws):
    v = cell.value
    if cell.data_type == "f" or (isinstance(v, str) and v.startswith("=")):
        cached = vws[cell.coordinate].value
        if cached is not None:
            return cached
    return v


def apply_json_to_xlsx(xlsx: str, jp: str) -> tuple[int, list[str]]:
    with open(jp) as f:
        data = json.load(f)
    lookup = {
        int(e["enumber"]): {"str": e.get("enumber_str", str(e["enumber"])),
                             "proposed": int(e["proposed"])}
        for e in data.get("mapping", []) if e.get("enumber") is not None
    }

    wb, wbv = openpyxl.load_workbook(xlsx), openpyxl.load_workbook(xlsx, data_only=True)
    ws, wsv = wb["input"], wbv["input"]
    fc, rc  = _col(ws, "found"), _col(ws, "replace")

    changed, log = 0, []
    for row in ws.iter_rows(min_row=2):
        fv = _eff(row[fc - 1], wsv)
        if not fv:
            continue
        m = ENUMBER_RE.search(Path(str(fv).strip()).stem)
        if not m:
            continue
        ei = int(m.group(1))
        if ei not in lookup or lookup[ei]["proposed"] == ei:
            continue
        prop = lookup[ei]["proposed"]
        rv   = _eff(row[rc - 1], wsv)
        if not rv:
            continue
        p   = Path(str(rv).strip())
        m2  = ENUMBER_RE.search(p.stem)
        if not m2:
            log.append(f"  [SKIP] no enumber in replace: {p.name}")
            continue
        new_str  = str(prop).zfill(len(m2.group(1)))
        new_stem = p.stem[: m2.start(1)] + new_str + p.stem[m2.end(1):]
        new_path = str(p.parent / (new_stem + p.suffix))
        row[rc - 1].value = new_path
        log.append(f"  {p.name}  →  {Path(new_path).name}")
        changed += 1

    wb.save(xlsx)
    return changed, log


def run_renames(xlsx: str, mode: str) -> tuple[int, list[str]]:
    wb, wbv = openpyxl.load_workbook(xlsx), openpyxl.load_workbook(xlsx, data_only=True)
    ws, wsv = wb["input"], wbv["input"]
    fc, rc  = _col(ws, "found"), _col(ws, "replace")
    count, log = 0, []
    for row in ws.iter_rows(min_row=2):
        src = _eff(row[fc - 1], wsv)
        dst = _eff(row[rc - 1], wsv)
        if not src or not dst:
            continue
        src, dst = str(src).strip(), str(dst).strip()
        if os.path.normcase(os.path.normpath(src)) == os.path.normcase(os.path.normpath(dst)):
            log.append(f"  [NO-OP]   {Path(src).name}")
            continue
        if not os.path.isfile(src):
            log.append(f"  [MISSING] {src}")
            continue
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        (shutil.copy2 if mode == "copy" else shutil.move)(src, dst)
        log.append(f"  {Path(src).name}  →  {Path(dst).name}")
        count += 1
    return count, log


# ── UI ────────────────────────────────────────────────────────────────────────

st.title("Camera Mapper")

# ─────────────────────────────────────────────────────────────────────────────
# Scan
# ─────────────────────────────────────────────────────────────────────────────
folder = st.text_input("Image folder:", DEFAULT_FOLDER)

if st.button("🔍 Scan"):
    with st.spinner("Scanning…"):
        images = scan_folder(folder)
    _xlsx = str(Path(folder) / "rename_files.xlsx")
    st.session_state.update({
        "cam_images":       images,
        "cam_folder":       folder,
        "click_seq":        [],
        "non_gps_proposed": {},
        "xlsx_path":        _xlsx,
        "json_path":        str(Path(folder) / "cam_mapping.json"),
    })
    st.session_state.pop("nogps_editor", None)

if "cam_images" not in st.session_state:
    st.stop()

images       = st.session_state["cam_images"]
cam_folder   = st.session_state["cam_folder"]
seq          = st.session_state.setdefault("click_seq", [])
non_gps_p    = st.session_state.setdefault("non_gps_proposed", {})
gps_images   = [r for r in images if r["lat"] is not None]
nogps_images = [r for r in images if r["lat"] is None]

# ── Diagnostics ───────────────────────────────────────────────────────────────
c1, c2, c3 = st.columns(3)
c1.metric("_INSTALL images found", len(images))
c2.metric("With GPS", len(gps_images))
c3.metric("Without GPS (table only)", len(nogps_images))

if not images:
    st.warning("No *_INSTALL* images found. Check the folder path and try again.")
    st.stop()

if not gps_images:
    st.warning(
        "No GPS EXIF data found in any image — map unavailable. "
        "Use the manual table below to assign proposed numbers."
    )

# ─────────────────────────────────────────────────────────────────────────────
# Step 1: Map + sequence assignment (GPS cameras only)
# ─────────────────────────────────────────────────────────────────────────────
if gps_images:
    st.markdown("### Step 1 — Build the camera sequence")
    st.info(
        "**The map is view-only** — clicking pins does nothing here.  \n"
        "**To assign the sequence:**  \n"
        "1. Hover a pin to see the photo.  \n"
        "2. Click **📋 Get List** (bottom-right corner of the map).  \n"
        "3. Copy the numbers shown in the popup.  \n"
        "4. Paste them into the box below, **reorder** so the first number is the new Cam #1.  \n"
        "5. Click **✅ Apply sequence**."
    )

    try:
        fmap = build_map(images, seq)
        if fmap is not None:
            map_html = _inject_get_list_button(fmap._repr_html_(), fmap.get_name())
            components.html(map_html, height=650, scrolling=False)
    except Exception as _exc:
        st.error(f"Map error: {_exc}")

    # ── Paste & reorder sequence ──────────────────────────────────────────────
    paste_seq = st.text_input(
        "📋 Paste camera numbers in new order (first = new Cam #1):",
        key="paste_seq_input",
        placeholder="e.g. 3,1,5,2,4",
        help="Click 'Get List' on the map → Copy → paste here → reorder → Apply.",
    )
    if st.button("✅ Apply sequence", key="apply_paste_btn"):
        raw = paste_seq.strip()
        if raw:
            try:
                pasted = [int(x.strip()) for x in raw.split(",") if x.strip()]
                enum_to_file = {r["enumber"]: r["file"] for r in gps_images}
                bad = [n for n in pasted if n not in enum_to_file]
                if bad:
                    st.warning(f"Unknown camera number(s): {bad}  — check the map pins.")
                else:
                    seq.clear()
                    seq.extend(enum_to_file[n] for n in pasted)
                    st.rerun()
            except ValueError:
                st.warning("Enter comma-separated integers, e.g.  3,1,5,2,4")
        else:
            st.warning("Paste numbers first.")

    # ── Manual add (one-at-a-time fallback) ──────────────────────────────────
    unassigned = [r for r in gps_images if r["file"] not in seq]
    if unassigned:
        with st.expander("Or add cameras one at a time", expanded=not seq):
            options = ["— pick a camera —"] + [
                f'Cam {r["enumber"]}  —  {r["file"]}' for r in unassigned
            ]
            pick = st.selectbox("Camera:", options, key="manual_add_sel")
            if pick != options[0]:
                if st.button("➕ Add to sequence", key="manual_add_btn"):
                    fname = unassigned[options.index(pick) - 1]["file"]
                    if fname not in seq:
                        seq.append(fname)
                    st.rerun()

    # ── Sequence display + controls ───────────────────────────────────────────
    n_assigned = len(seq)
    n_gps      = len(gps_images)
    if n_assigned:
        st.markdown(f"**Sequence — {n_assigned} of {n_gps} GPS camera(s) assigned:**")
        cols = st.columns(min(n_assigned, 4))
        for i, fname in enumerate(seq):
            cols[i % len(cols)].markdown(f"`{i+1}.` {fname}")

        bc1, bc2, _ = st.columns([1, 1, 4])
        with bc1:
            if st.button("↩ Undo last", key="undo_btn"):
                seq.pop()
                st.rerun()
        with bc2:
            if st.button("✕ Clear all", key="clear_btn"):
                seq.clear()
                st.rerun()
    else:
        st.info(
            "Use the map **Get List** → paste + reorder above, "
            "or add cameras one at a time via the expander."
        )

# ─────────────────────────────────────────────────────────────────────────────
# Step 2: Manual table for non-GPS cameras
# ─────────────────────────────────────────────────────────────────────────────
if nogps_images:
    st.markdown("### Step 2 — Assign numbers for cameras without GPS")
    ng_df = pd.DataFrame([
        {
            "File":       r["file"],
            "Current #":  r["enumber"],
            "Proposed #": non_gps_p.get(r["file"], r["enumber"]),
        }
        for r in nogps_images
    ])
    edited_ng = st.data_editor(
        ng_df,
        column_config={
            "File":       st.column_config.TextColumn("File"),
            "Current #":  st.column_config.NumberColumn("Current #"),
            "Proposed #": st.column_config.NumberColumn("Proposed #", required=True),
        },
        disabled=["File", "Current #"],
        hide_index=True,
        use_container_width=True,
        key="nogps_editor",
    )
    for _, row in edited_ng.iterrows():
        if pd.notna(row["Proposed #"]):
            non_gps_p[row["File"]] = int(row["Proposed #"])

# ─────────────────────────────────────────────────────────────────────────────
# Mapping summary (read-only)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("### Mapping summary")
seq_pos = {f: i + 1 for i, f in enumerate(seq)}
summary = []
for r in images:
    if r["lat"] is not None:
        proposed = seq_pos.get(r["file"], "— not yet assigned —")
    else:
        proposed = non_gps_p.get(r["file"], r["enumber"])
    summary.append({
        "File":       r["file"],
        "Current #":  r["enumber"],
        "Proposed #": proposed,
        "GPS":        "✓" if r["lat"] is not None else "—",
    })
st.dataframe(
    pd.DataFrame(summary),
    hide_index=True,
    use_container_width=True,
    column_config={"Proposed #": st.column_config.TextColumn("Proposed #")},
)

# ─────────────────────────────────────────────────────────────────────────────
# Save JSON
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
xlsx_path = st.text_input(
    "rename_files.xlsx path (cam_mapping.json saves to the same folder):",
    value=str(Path(cam_folder) / "rename_files.xlsx"),
    key="xlsx_path_input",
)

pc, sc = st.columns(2)
with pc:
    if st.button("👁 Preview JSON", key="preview_btn"):
        sp   = {f: i + 1 for i, f in enumerate(seq)}
        prev = {
            "folder":    cam_folder,
            "generated": "…",
            "mapping": [
                {
                    "file":        r["file"],
                    "enumber":     r["enumber"],
                    "enumber_str": r["enumber_str"],
                    "proposed":    sp.get(r["file"], non_gps_p.get(r["file"], r["enumber"])),
                }
                for r in images if r["enumber"] is not None
            ],
        }
        st.json(prev)

with sc:
    if st.button("💾 Save cam_mapping.json", key="save_json_btn"):
        if not xlsx_path:
            st.error("Enter the xlsx path above first.")
        else:
            try:
                out = save_json(xlsx_path, cam_folder, images, seq, non_gps_p)
                st.session_state.update({"json_path": out, "xlsx_path": xlsx_path})
                st.success(f"Saved → {out}")
            except Exception as exc:
                st.error(str(exc))

# ─────────────────────────────────────────────────────────────────────────────
# Apply JSON → Excel
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
with st.expander("📋  Apply JSON → update Excel 'replace' column", expanded=False):
    st.caption(
        "Reads cam_mapping.json and rewrites the 'replace' column of the 'input' sheet "
        "so each filename uses the proposed camera number (zero-padding preserved)."
    )
    xl1 = st.text_input("rename_files.xlsx:", st.session_state.get("xlsx_path", ""), key="xl_apply")
    jf1 = st.text_input(
        "cam_mapping.json:",
        value=st.session_state.get("json_path", _json_path(xl1)),
        key="jf_apply",
    )
    if st.button("📝 Apply JSON to Excel", key="apply_json_btn"):
        ok = True
        if not xl1 or not os.path.isfile(xl1):
            st.error("XLSX not found.")
            ok = False
        if not jf1 or not os.path.isfile(jf1):
            st.error("JSON not found.")
            ok = False
        if ok:
            try:
                n, log = apply_json_to_xlsx(xl1, jf1)
                if log:
                    st.code("\n".join(log))
                st.success(f"Updated {n} row(s) in 'input' sheet — saved.")
            except Exception as exc:
                st.error(str(exc))

# ─────────────────────────────────────────────────────────────────────────────
# Apply renames (copy / move)
# ─────────────────────────────────────────────────────────────────────────────
st.divider()
with st.expander("🚀  Apply renames (copy / move files)", expanded=False):
    st.caption("Run this after applying the JSON above. Reads the current 'input' sheet as-is.")
    xl2  = st.text_input("rename_files.xlsx:", st.session_state.get("xlsx_path", ""), key="xl_rename")
    mode = st.radio("Mode", ["copy", "move"], horizontal=True, key="rename_mode")
    if mode == "move":
        st.warning("**Move** permanently relocates files. Verify the 'replace' column is correct first.")
        confirmed = st.checkbox("Yes, I've verified — proceed with move.", key="move_confirm")
    else:
        confirmed = True
    if st.button("▶️ Apply renames", key="apply_rename_btn", disabled=not confirmed):
        if not xl2 or not os.path.isfile(xl2):
            st.error("XLSX not found.")
        else:
            try:
                n, log = run_renames(xl2, mode)
                if log:
                    st.code("\n".join(log))
                st.success(f"Done — {n} file(s) {mode}d.")
            except Exception as exc:
                st.error(str(exc))
