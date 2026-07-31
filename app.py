"""
app.py  —  Camera Mapper (Flask)

Run:  python app.py
Then open http://localhost:5000 in your browser.

Workflow:
  1. Enter image folder path and click Scan.
  2. Click camera pins on the satellite map in the desired new-number order.
     Pins turn green with the proposed number as you assign them.
  3. Save cam_mapping.json (same folder as rename_files.xlsx).
  4. Apply JSON  → rewrites 'replace' column in rename_files.xlsx input sheet.
  5. Run copy / move from the app.
"""
from __future__ import annotations

import base64, json, os, re, shutil
from datetime import datetime
from io import BytesIO
from pathlib import Path

import openpyxl
import piexif
from flask import Flask, jsonify, render_template, request, send_file
from PIL import Image


def _resolve_share_root() -> Path:
    """Resolve the LAUSD OneDrive share root dynamically (mirrors the
    convention in lausd_tools/acro_form_fill_batch.py). Override with
    LAUSD_SHARE_ROOT if the default candidates don't match."""
    env_override = os.environ.get("LAUSD_SHARE_ROOT", "").strip()
    if env_override:
        return Path(env_override)
    candidates = [
        Path.home() / "OneDrive - Los Angeles Unified School District" / "Paul's Share Folder",
        Path.home() / "OneDrive" / "Paul's Share Folder",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return candidates[0]

# ── app setup ─────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = "cam_mapper_local"

# Single-user in-memory state (local app only)
STATE: dict = {
    "images":           [],   # list of dicts from scan_folder()
    "folder":           "",
    "sequence":         [],   # filenames in assigned order
    "non_gps_proposed": {},   # filename → proposed int
}

# ── constants ─────────────────────────────────────────────────────────────────
INSTALL_RE = re.compile(r"_INSTALL", re.IGNORECASE)
ENUMBER_RE = re.compile(r"(\d+)((?:_[A-Za-z]+|[A-Za-z]+)*)$")
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".tif", ".tiff"}

# lausd_tools project files (sibling directory)
_LAUSD_TOOLS_DIR  = Path(__file__).parent.parent / "lausd_tools"
_SCHOOLS_JSON     = _LAUSD_TOOLS_DIR / "temp_r1.json"
_TEMPLATES_JSON   = _LAUSD_TOOLS_DIR / "visio_templates.json"
_OUTPUT_DIR_FALLBACK = str(_resolve_share_root())
_CAM_SUBPATH = "Camera/Design/Pictures"
_XLSX_NAME   = "rename_files.xlsx"

# ── school resolver (mirrors rename_files.py / image_cleaner.py logic) ────────

def _sanitize(name: str) -> str:
    return "".join("_" if c in '<>:"/\\|?*' else c for c in name).strip() or "Unknown"


def _remap_legacy_user_path(path: str) -> str:
    """visio_templates.json stores paths under the original author's
    C:/Users/Paul profile; remap that prefix onto the current user's home."""
    normalized = path.replace("\\", "/")
    legacy_prefix = "C:/Users/Paul/"
    if not normalized.startswith(legacy_prefix):
        return path
    share_marker = "OneDrive - Los Angeles Unified School District/Paul's Share Folder"
    if share_marker in normalized:
        tail = normalized.split(share_marker, 1)[1].lstrip("/")
        return str(_resolve_share_root() / Path(tail))
    return str(Path.home() / Path(normalized[len(legacy_prefix):]))


def _output_dir() -> Path:
    if _TEMPLATES_JSON.exists():
        try:
            jobs = json.loads(_TEMPLATES_JSON.read_text(encoding="utf-8"))
            if isinstance(jobs, list) and jobs and "OUTPUT_DIR" in jobs[0]:
                return Path(_remap_legacy_user_path(jobs[0]["OUTPUT_DIR"]))
        except Exception:
            pass
    return Path(_OUTPUT_DIR_FALLBACK)


def _active_schools() -> list[dict]:
    """Return [{label, folder, xlsx_path, json_path}] for every active school."""
    if not _SCHOOLS_JSON.exists():
        return []
    try:
        schools = json.loads(_SCHOOLS_JSON.read_text(encoding="utf-8"))
    except Exception:
        return []

    base = _output_dir()
    result = []
    for s in schools:
        if str(s.get("activate") or "").strip().lower() != "x":
            continue
        site = str(s.get("Site") or s.get("School Name") or "Unknown").strip()
        loc  = s.get("Loc Code")
        if loc is None:
            loc_str = ""
        elif isinstance(loc, float) and loc.is_integer():
            loc_str = str(int(loc))
        else:
            loc_str = str(loc).strip()

        suffix  = f" ({loc_str})" if loc_str else ""
        proj    = base / "Project Sites" / f"{_sanitize(site)}{suffix}"
        folder  = proj / Path(_CAM_SUBPATH)
        result.append({
            "label":     f"{site}{suffix}",
            "folder":    str(folder),
            "xlsx_path": str(folder / _XLSX_NAME),
            "json_path": str(folder / "cam_mapping.json"),
        })
    return result


# ── helpers ───────────────────────────────────────────────────────────────────

# Private, non-standard EXIF tag (Exif IFD) used to stash the camera's Visio
# icon shape type ("default" / "180" / "270" — how many of its 4 lens heads
# to draw). No standard EXIF field covers this; a private tag number avoids
# colliding with any real field (e.g. ImageDescription/UserComment) that
# other tools or the camera itself might legitimately write.
SHAPE_TYPE_EXIF_TAG = 0xC7A1

# piexif.dump() raises KeyError on any tag id it doesn't already know the
# type of — register ours (ASCII) so writes to the Exif IFD succeed.
piexif.TAGS["Exif"][SHAPE_TYPE_EXIF_TAG] = {"name": "ShapeType", "type": piexif.TYPES.Ascii}

def _get_gps_exif(path: str):
    try:
        img = Image.open(path)
        exif_bytes = img.info.get("exif", b"")
        img.close()
        if not exif_bytes:
            return None, None, None, None
        exif_dict = piexif.load(exif_bytes)
        gps = exif_dict.get("GPS", {})
        exif_ifd = exif_dict.get("Exif", {})
        shape_type = None
        raw_shape_type = exif_ifd.get(SHAPE_TYPE_EXIF_TAG)
        if raw_shape_type:
            try:
                shape_type = raw_shape_type.decode("ascii")
            except Exception:
                pass
        if not gps:
            return None, None, None, shape_type
        def _deg(v):
            a, b, c = v
            return a[0] / a[1] + b[0] / b[1] / 60 + c[0] / c[1] / 3600
        lat  = gps.get(piexif.GPSIFD.GPSLatitude)
        lref = gps.get(piexif.GPSIFD.GPSLatitudeRef)
        lon  = gps.get(piexif.GPSIFD.GPSLongitude)
        lnrf = gps.get(piexif.GPSIFD.GPSLongitudeRef)
        heading = None
        try:
            num, den = gps[piexif.GPSIFD.GPSImgDirection]
            if den:
                heading = num / den
        except Exception:
            pass
        if lat and lref and lon and lnrf:
            return (
                _deg(lat) * (-1 if lref == b"S" else 1),
                _deg(lon) * (-1 if lnrf == b"W" else 1),
                heading,
                shape_type,
            )
    except Exception:
        pass
    return None, None, None, None


def _enumber(stem: str):
    m = ENUMBER_RE.search(stem)
    return (int(m.group(1)), m.group(1), m.group(2)) if m else (None, None, None)


def _col(ws, name: str) -> int:
    for c in ws[1]:
        if c.value and str(c.value).strip().lower() == name.lower():
            return c.column
    raise ValueError(f"Column '{name}' not found in sheet '{ws.title}'")


def _get_or_add_col(ws, name: str) -> int:
    for c in ws[1]:
        if c.value and str(c.value).strip().lower() == name.lower():
            return c.column
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=name)
    return new_col


def _eff(cell, vws):
    v = cell.value
    if cell.data_type == "f" or (isinstance(v, str) and v.startswith("=")):
        cached = vws[cell.coordinate].value
        if cached is not None:
            return cached
    return v


def _seq_state() -> list[dict]:
    """Return serializable sequence state for all GPS cameras."""
    seq_pos = {f: i + 1 for i, f in enumerate(STATE["sequence"])}
    return [
        {
            "file":     r["file"],
            "enumber":  r["enumber"],
            "proposed": seq_pos.get(r["file"]),
        }
        for r in STATE["images"] if r["lat"] is not None
    ]


# ── routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/schools")
def schools():
    return jsonify({"schools": _active_schools()})


@app.route("/scan", methods=["POST"])
def scan():
    folder = (request.json or {}).get("folder", "").strip()
    if not os.path.isdir(folder):
        return jsonify({"error": f"Folder not found: {folder}"}), 400

    images = []
    try:
        entries = sorted(Path(folder).iterdir(), key=lambda e: e.name.lower())
    except OSError as exc:
        return jsonify({"error": str(exc)}), 400

    for e in entries:
        if not e.is_file() or e.suffix.lower() not in IMAGE_EXTS:
            continue
        if not INSTALL_RE.search(e.stem):
            continue
        enum, estr, esuf = _enumber(e.stem)
        lat, lon, heading, shape_type = _get_gps_exif(str(e))
        images.append({
            "path":        str(e),
            "file":        e.name,
            "enumber":     enum,
            "enumber_str": estr,
            "esuffix":     esuf,
            "lat":         lat,
            "lon":         lon,
            "heading":     heading,
            "shape_type":  shape_type,
        })

    STATE.update({"images": images, "folder": folder,
                  "sequence": [], "non_gps_proposed": {}})

    return jsonify({
        "images":    images,
        "folder":    folder,
        "xlsx_path": str(Path(folder) / "rename_files.xlsx"),
        "json_path": str(Path(folder) / "cam_mapping.json"),
    })


@app.route("/sequence/add", methods=["POST"])
def sequence_add():
    enumber = (request.json or {}).get("enumber")
    img = next((r for r in STATE["images"] if r["enumber"] == enumber), None)
    if not img:
        return jsonify({"error": f"Camera {enumber} not found"}), 404
    if img["file"] not in STATE["sequence"]:
        STATE["sequence"].append(img["file"])
    return jsonify({"sequence": _seq_state()})


@app.route("/sequence/remove", methods=["POST"])
def sequence_remove():
    enumber = (request.json or {}).get("enumber")
    img = next((r for r in STATE["images"] if r["enumber"] == enumber), None)
    if img and img["file"] in STATE["sequence"]:
        STATE["sequence"].remove(img["file"])
    return jsonify({"sequence": _seq_state()})


@app.route("/sequence/undo", methods=["POST"])
def sequence_undo():
    if STATE["sequence"]:
        STATE["sequence"].pop()
    return jsonify({"sequence": _seq_state()})


@app.route("/sequence/clear", methods=["POST"])
def sequence_clear():
    STATE["sequence"] = []
    return jsonify({"sequence": _seq_state()})


@app.route("/non_gps", methods=["POST"])
def non_gps():
    for fname, val in ((request.json or {}).get("updates", {})).items():
        try:
            STATE["non_gps_proposed"][fname] = int(val)
        except (ValueError, TypeError):
            pass
    return jsonify({"ok": True})


@app.route("/thumbnail")
def thumbnail():
    path = request.args.get("path", "")
    if not path or not os.path.isfile(path):
        return "", 404
    # Verify the file is within the scanned folder
    try:
        Path(path).resolve().relative_to(Path(STATE["folder"]).resolve())
    except ValueError:
        return "", 403
    try:
        img = Image.open(path)
        img.thumbnail((320, 320))
        buf = BytesIO()
        img.save(buf, "JPEG", quality=82)
        img.close()
        buf.seek(0)
        return send_file(buf, mimetype="image/jpeg")
    except Exception:
        return "", 500


@app.route("/save_json", methods=["POST"])
def save_json():
    body     = request.json or {}
    xlsx     = body.get("xlsx", "").strip()
    start    = max(1, int(body.get("start", 1) or 1))
    filename = re.sub(r'[^\w\-. ]', '_', body.get("filename", "cam_mapping").strip() or "cam_mapping")
    if not xlsx:
        return jsonify({"error": "xlsx path required"}), 400

    seq_pos   = {f: i + start for i, f in enumerate(STATE["sequence"])}
    seq_files = set(STATE["sequence"])
    mapping   = []
    for r in STATE["images"]:
        if r["enumber"] is None:
            continue
        if r["lat"] is not None:
            proposed = seq_pos.get(r["file"], r["enumber"])
        else:
            proposed = STATE["non_gps_proposed"].get(r["file"], r["enumber"])
        mapping.append({
            "file":        r["file"],
            "enumber":     r["enumber"],
            "enumber_str": r["enumber_str"],
            "proposed":    proposed,
            "sequenced":   r["file"] in seq_files,
            "lat":         r.get("lat"),
            "lon":         r.get("lon"),
            "heading":     r.get("heading"),
            "shape_type":  r.get("shape_type"),
        })

    out = str(Path(xlsx).parent / f"{filename}.json")
    Path(out).write_text(json.dumps({
        "folder":    STATE["folder"],
        "generated": datetime.now().isoformat(timespec="seconds"),
        "mapping":   mapping,
    }, indent=2), encoding="utf-8")
    return jsonify({"path": out})


def _write_gps_exif(path: str, lat: float, lon: float):
    if Path(path).suffix.lower() not in {".jpg", ".jpeg"}:
        raise ValueError(f"GPS EXIF write only supported for JPEG files")

    def _dms(v):
        d = int(v)
        m = int((v - d) * 60)
        s = round((v - d - m / 60) * 3600 * 100)   # hundredths of a second
        return ((d, 1), (m, 1), (s, 100))

    try:
        exif_dict = piexif.load(path)
    except Exception:
        exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "Interop": {}, "1st": {}}

    exif_dict.setdefault("GPS", {}).update({
        piexif.GPSIFD.GPSLatitudeRef:  ("N" if lat >= 0 else "S").encode(),
        piexif.GPSIFD.GPSLatitude:     _dms(abs(lat)),
        piexif.GPSIFD.GPSLongitudeRef: ("E" if lon >= 0 else "W").encode(),
        piexif.GPSIFD.GPSLongitude:    _dms(abs(lon)),
    })
    piexif.insert(piexif.dump(exif_dict), path)


def _write_heading_exif(path: str, heading: float):
    if Path(path).suffix.lower() not in {".jpg", ".jpeg"}:
        raise ValueError(f"GPS EXIF write only supported for JPEG files")

    try:
        exif_dict = piexif.load(path)
    except Exception:
        exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "Interop": {}, "1st": {}}

    heading = heading % 360
    exif_dict.setdefault("GPS", {}).update({
        piexif.GPSIFD.GPSImgDirectionRef: b"T",
        piexif.GPSIFD.GPSImgDirection:    (round(heading * 100), 100),
    })
    piexif.insert(piexif.dump(exif_dict), path)


def _write_shape_type_exif(path: str, shape_type: str):
    if Path(path).suffix.lower() not in {".jpg", ".jpeg"}:
        raise ValueError(f"GPS EXIF write only supported for JPEG files")

    try:
        exif_dict = piexif.load(path)
    except Exception:
        exif_dict = {"0th": {}, "Exif": {}, "GPS": {}, "Interop": {}, "1st": {}}

    exif_dict.setdefault("Exif", {})[SHAPE_TYPE_EXIF_TAG] = shape_type.encode("ascii")
    piexif.insert(piexif.dump(exif_dict), path)


@app.route("/place_camera", methods=["POST"])
def place_camera():
    body = request.json or {}
    file = body.get("file", "").strip()
    lat  = body.get("lat")
    lon  = body.get("lon")

    if not file or lat is None or lon is None:
        return jsonify({"error": "file, lat, lon required"}), 400

    img = next((r for r in STATE["images"] if r["file"] == file), None)
    if not img:
        return jsonify({"error": f"File not found in scan: {file}"}), 404

    img["lat"] = float(lat)
    img["lon"] = float(lon)

    return jsonify({
        "file":    img["file"],
        "enumber": img["enumber"],
        "lat":     img["lat"],
        "lon":     img["lon"],
        "path":    img["path"],
    })


@app.route("/write_gps", methods=["POST"])
def write_gps():
    body = request.json or {}
    file = body.get("file", "").strip()
    lat  = body.get("lat")
    lon  = body.get("lon")

    if not file or lat is None or lon is None:
        return jsonify({"error": "file, lat, lon required"}), 400

    img = next((r for r in STATE["images"] if r["file"] == file), None)
    if not img:
        return jsonify({"error": f"File not in scan: {file}"}), 404

    path = img["path"]
    if not os.path.isfile(path):
        return jsonify({"error": f"Image not found on disk: {path}"}), 404

    try:
        _write_gps_exif(path, float(lat), float(lon))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"ok": True})


@app.route("/set_heading", methods=["POST"])
def set_heading():
    body    = request.json or {}
    file    = body.get("file", "").strip()
    heading = body.get("heading")

    if not file or heading is None:
        return jsonify({"error": "file, heading required"}), 400

    img = next((r for r in STATE["images"] if r["file"] == file), None)
    if not img:
        return jsonify({"error": f"File not found in scan: {file}"}), 404

    img["heading"] = float(heading) % 360

    return jsonify({
        "file":    img["file"],
        "enumber": img["enumber"],
        "heading": img["heading"],
    })


@app.route("/write_heading", methods=["POST"])
def write_heading():
    body    = request.json or {}
    file    = body.get("file", "").strip()
    heading = body.get("heading")

    if not file or heading is None:
        return jsonify({"error": "file, heading required"}), 400

    img = next((r for r in STATE["images"] if r["file"] == file), None)
    if not img:
        return jsonify({"error": f"File not in scan: {file}"}), 404

    path = img["path"]
    if not os.path.isfile(path):
        return jsonify({"error": f"Image not found on disk: {path}"}), 404

    try:
        _write_heading_exif(path, float(heading))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"ok": True})


SHAPE_TYPES = {"default", "180", "270"}


@app.route("/set_shape_type", methods=["POST"])
def set_shape_type():
    body       = request.json or {}
    file       = body.get("file", "").strip()
    shape_type = body.get("shape_type")

    if not file or shape_type not in SHAPE_TYPES:
        return jsonify({"error": f"file required, shape_type must be one of {sorted(SHAPE_TYPES)}"}), 400

    img = next((r for r in STATE["images"] if r["file"] == file), None)
    if not img:
        return jsonify({"error": f"File not found in scan: {file}"}), 404

    img["shape_type"] = shape_type

    return jsonify({
        "file":       img["file"],
        "enumber":    img["enumber"],
        "shape_type": img["shape_type"],
    })


@app.route("/write_shape_type", methods=["POST"])
def write_shape_type():
    body       = request.json or {}
    file       = body.get("file", "").strip()
    shape_type = body.get("shape_type")

    if not file or shape_type not in SHAPE_TYPES:
        return jsonify({"error": f"file required, shape_type must be one of {sorted(SHAPE_TYPES)}"}), 400

    img = next((r for r in STATE["images"] if r["file"] == file), None)
    if not img:
        return jsonify({"error": f"File not in scan: {file}"}), 404

    path = img["path"]
    if not os.path.isfile(path):
        return jsonify({"error": f"Image not found on disk: {path}"}), 404

    try:
        _write_shape_type_exif(path, shape_type)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"ok": True})


@app.route("/save_map_image", methods=["POST"])
def save_map_image():
    body    = request.json or {}
    img_b64 = body.get("image", "")
    if not img_b64:
        return jsonify({"error": "No image data"}), 400

    folder = body.get("folder", "").strip() or STATE.get("folder", "")
    if not folder:
        return jsonify({"error": "No project folder available — scan a folder first or load a mapping"}), 400

    # …/Camera/Design/Pictures  →  …/Camera/Design/Camera Layout/floor_plans
    floor_plans = Path(folder).parent / "Camera Layout" / "floor_plans"
    floor_plans.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = floor_plans / f"map_{timestamp}.png"

    try:
        out_path.write_bytes(base64.b64decode(img_b64))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"path": str(out_path)})


@app.route("/load_mapping", methods=["POST"])
def load_mapping():
    jp = (request.json or {}).get("json_path", "").strip()
    if not jp:
        return jsonify({"error": "json_path required"}), 400
    if not os.path.isfile(jp):
        return jsonify({"error": f"File not found: {jp}"}), 404
    try:
        data = json.loads(Path(jp).read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify({"error": f"Could not parse JSON: {e}"}), 400

    mapping = data.get("mapping", [])
    folder  = data.get("folder", "")

    # If any entry lacks lat/lon (JSON saved before coordinates were recorded),
    # fall back to reading GPS directly from the image files in the folder.
    needs_gps = any(e.get("lat") is None for e in mapping)
    if needs_gps and folder and os.path.isdir(folder):
        file_paths = {
            p.name: str(p)
            for p in Path(folder).iterdir()
            if INSTALL_RE.search(p.name) and p.suffix.lower() in IMAGE_EXTS
        }
        for entry in mapping:
            if entry.get("lat") is not None:
                continue
            path = file_paths.get(entry.get("file", ""))
            if not path:
                continue
            lat, lon, heading, shape_type = _get_gps_exif(path)
            if lat is not None:
                entry["lat"] = lat
                entry["lon"] = lon
            if entry.get("heading") is None:
                entry["heading"] = heading
            if entry.get("shape_type") is None:
                entry["shape_type"] = shape_type

    return jsonify({
        "folder":    folder,
        "generated": data.get("generated", ""),
        "mapping":   mapping,
    })


@app.route("/apply_json", methods=["POST"])
def apply_json():
    xlsx = (request.json or {}).get("xlsx", "").strip()
    jp   = (request.json or {}).get("json_path", "").strip()

    if not os.path.isfile(xlsx):
        return jsonify({"error": f"XLSX not found: {xlsx}"}), 400
    if not os.path.isfile(jp):
        return jsonify({"error": f"JSON not found: {jp}"}), 400

    with open(jp, encoding="utf-8") as f:
        data = json.load(f)

    # enumber → proposed number for every camera in the JSON
    lookup = {
        int(e["enumber"]): int(e["proposed"])
        for e in data.get("mapping", [])
        if e.get("enumber") is not None
    }
    # Cameras the user actually clicked into the sequence (sequenced flag,
    # with fallback to "number changed" for JSON files saved before this field existed)
    selected = {
        int(e["enumber"])
        for e in data.get("mapping", [])
        if e.get("enumber") is not None
        and (e.get("sequenced") or int(e["enumber"]) != int(e.get("proposed", e["enumber"])))
    }

    wb  = openpyxl.load_workbook(xlsx)
    wbv = openpyxl.load_workbook(xlsx, data_only=True)
    ws, wsv = wb["input"], wbv["input"]
    fc, rc  = _col(ws, "found"), _col(ws, "replace")
    sc      = _get_or_add_col(ws, "selected")

    # Optional columns — skip gracefully if not present
    def _try_col(name):
        try:
            return _col(ws, name)
        except ValueError:
            return None

    nc = _try_col("number")   # proposed camera number
    xc = _try_col("suffix")   # alpha suffix (e.g. _INSTALL)

    # Clear managed columns so stale values from prior runs don't linger
    for row in ws.iter_rows(min_row=2):
        row[sc - 1].value = None
        if nc: row[nc - 1].value = None
        if xc: row[xc - 1].value = None

    changed, log = 0, []
    for row in ws.iter_rows(min_row=2):
        fv = _eff(row[fc - 1], wsv)
        if not fv:
            continue
        m = ENUMBER_RE.search(Path(str(fv).strip()).stem)
        if not m:
            continue
        ei      = int(m.group(1))
        esuffix = m.group(2)   # e.g. "_INSTALL" or ""

        if ei not in selected:
            continue

        # Mark row as selected
        row[sc - 1].value = "x"

        # Always write number + suffix for sequenced cameras
        proposed = lookup.get(ei, ei)
        if nc:
            row[nc - 1].value = proposed
        if xc:
            row[xc - 1].value = esuffix

        # Update replace path only when the number actually changed
        if lookup.get(ei, ei) == ei:
            continue
        rv = _eff(row[rc - 1], wsv)
        if not rv:
            continue
        p  = Path(str(rv).strip())
        m2 = ENUMBER_RE.search(p.stem)
        if not m2:
            log.append(f"[SKIP] no enumber in replace path: {p.name}")
            continue
        new_str  = str(proposed).zfill(len(m2.group(1)))
        new_stem = p.stem[: m2.start(1)] + new_str + p.stem[m2.end(1):]
        new_path = str(p.parent / (new_stem + p.suffix))
        row[rc - 1].value = new_path
        log.append(f"{p.name}  →  {Path(new_path).name}")
        changed += 1

    wb.save(xlsx)
    return jsonify({"changed": changed, "log": log})


@app.route("/run_renames", methods=["POST"])
def run_renames():
    xlsx = (request.json or {}).get("xlsx", "").strip()
    mode = (request.json or {}).get("mode", "copy")

    if mode not in ("copy", "move"):
        return jsonify({"error": "mode must be 'copy' or 'move'"}), 400
    if not os.path.isfile(xlsx):
        return jsonify({"error": f"XLSX not found: {xlsx}"}), 400

    wb  = openpyxl.load_workbook(xlsx)
    wbv = openpyxl.load_workbook(xlsx, data_only=True)
    ws, wsv = wb["input"], wbv["input"]
    fc, rc  = _col(ws, "found"), _col(ws, "replace")

    count, log = 0, []
    for row in ws.iter_rows(min_row=2):
        src = _eff(row[fc - 1], wsv)
        dst = _eff(row[rc - 1], wsv)
        if not src or not dst:
            continue
        src, dst = str(src).strip(), str(dst).strip()
        if os.path.normcase(os.path.normpath(src)) == os.path.normcase(os.path.normpath(dst)):
            log.append(f"[NO-OP]   {Path(src).name}")
            continue
        if not os.path.isfile(src):
            log.append(f"[MISSING] {src}")
            continue
        os.makedirs(os.path.dirname(dst), exist_ok=True)
        (shutil.copy2 if mode == "copy" else shutil.move)(src, dst)
        log.append(f"{Path(src).name}  →  {Path(dst).name}")
        count += 1

    return jsonify({"count": count, "log": log})


# ── entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import threading, webbrowser
    threading.Timer(1.2, lambda: webbrowser.open("http://localhost:5000")).start()
    app.run(debug=False, port=5000)
